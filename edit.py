from openpyxl import Workbook
from openpyxl.styles import Border, Side

def init(wb, sheet_name):
    ws = wb.create_sheet(sheet_name)
    template_header_list = [
        "학년",
        "반",
        "학교명",
        "학급 인원",
        "국어",
        "수학",
        "영어",
        "한국사",
        "탐구",
        "실시일",
    ]

    write_list_to_excel(ws,2,2,template_header_list[:1])
    write_list_to_excel(ws,2,4,template_header_list[1:])

    add_border_to_range(ws,2,2,3,2)
    add_border_to_range(ws,2,4,3,12)

    template_content_title_list = [
        "국어 영역",
        "수학 영역",
        "영어 영역",
        "한국사 영역",
        "탐구 영역"
    ]

    write_list_to_excel(ws,5,5,template_content_title_list[:2],5)

    ws.merge_cells("E5:I5")
    ws.merge_cells("J5:N5")

    write_list_to_excel(ws,5,15,template_content_title_list[2:])
    ws.merge_cells("Q5:Z5")

    template_content_sub_title_list = [
        "번호",
        "성명",
        "선택 과목",
        "표준점수",
        "학교석차",
        "전국백분위",
        "등급"
    ]

    write_list_to_excel(ws,6,2,template_content_sub_title_list[:2],2)

    for i in range(2):
        write_list_to_excel(ws,6,5+i*(len(template_content_sub_title_list)-2),template_content_sub_title_list[2:])
    
    write_list_to_excel(ws,6,15,[template_content_sub_title_list[-1] for i in range(2)])
    write_list_to_excel(ws,6,17,template_content_sub_title_list[2:]*2)

    add_border_to_range(ws,5,2,30,2)
    add_border_to_range(ws,5,4,30,26)

def write_list_to_excel(ws, start_row, start_col, input_list, gap = 1):
    # 리스트 값들을 차례로 입력
    current_row = start_row
    current_col = start_col
    
    for value in input_list:
        ws.cell(row=current_row, column=current_col, value=value)
        current_col += gap

def add_border_to_range(ws, start_row, start_col, end_row, end_col):
    # Border 객체 생성
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # 특정 범위의 셀에 테두리 설정
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

def edit_header(wb, sheet_name, input_list):
    ws = wb.get_sheet_by_name(sheet_name)

    write_list_to_excel(ws,3,2,input_list[:1])
    write_list_to_excel(ws,3,4,input_list[1:])

def edit_content(wb, sheet_name, input_list):
    ws = wb.get_sheet_by_name(sheet_name)

    for i in range(len(input_list)):
        write_list_to_excel(ws,7+i,2,input_list[i][0:1])
        write_list_to_excel(ws,7+i,4,input_list[i][1:2])
        write_list_to_excel(ws,7+i,5,input_list[i][2])

def edit_footer(wb, sheet_name, list):
    print("here3")